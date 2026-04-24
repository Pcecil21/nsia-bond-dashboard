"""
Filename Suggester

Reads the first few pages of a document (PDF or text-like) and asks Claude
for a concise, descriptive filename. Used by the Inbox page when archiving
files from Ingestion/ to permanent NSIA folders — gives long-term archives
readable names instead of Docusign IDs and vendor-generated strings.

Output format convention: YYYY-MM__descriptive_topic.ext
    - Date prefix keeps files chronological in the folder
    - Double underscore separates date from topic (matches ingestion convention)
    - Topic is snake_case-ish, no spaces, no special chars

Non-PDF files (.docx, .xlsx, .png) fall back to a light cleanup of the
original name — we don't try to read their content here to keep it fast
and predictable.
"""

import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# Keep the Claude call fast + cheap — first few pages are usually enough
# to identify what a document is.
MAX_CHARS_FOR_PROMPT = 6000
MAX_PAGES_TO_READ = 3


def _read_pdf_head(pdf_path: Path) -> str:
    """Pull text from the first few pages of a PDF. Returns empty string on
    failure — caller handles fallback."""
    try:
        import fitz  # PyMuPDF, already in requirements.txt
    except ImportError:
        logger.warning("PyMuPDF not available — skipping content read")
        return ""

    try:
        with fitz.open(str(pdf_path)) as doc:
            pages_to_read = min(len(doc), MAX_PAGES_TO_READ)
            chunks = []
            for i in range(pages_to_read):
                text = doc[i].get_text()
                if text.strip():
                    chunks.append(text)
            full = "\n\n".join(chunks).strip()
            return full[:MAX_CHARS_FOR_PROMPT]
    except Exception as e:
        logger.warning("Failed to read %s: %s", pdf_path.name, e)
        return ""


def _read_xlsx_head(xlsx_path: Path) -> str:
    """Summarize an Excel file: sheet names + first ~30 rows of the first
    sheet. Enough context for Claude to identify what the file tracks
    (budget submission, reconciliation, ledger, etc)."""
    try:
        import openpyxl
    except ImportError:
        return ""
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        lines = [f"Sheet names: {', '.join(sheet_names)}", ""]

        first_sheet = wb[sheet_names[0]]
        lines.append(f"First 30 rows of '{sheet_names[0]}':")
        for i, row in enumerate(first_sheet.iter_rows(values_only=True)):
            if i >= 30:
                break
            # Keep non-empty cells, flatten to a compact pipe-delimited row
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append(" | ".join(cells))

        wb.close()
        text = "\n".join(lines)
        return text[:MAX_CHARS_FOR_PROMPT]
    except Exception as e:
        logger.warning("Failed to read xlsx %s: %s", xlsx_path.name, e)
        return ""


def _read_docx_head(docx_path: Path) -> str:
    """Pull the first ~40 paragraphs from a Word doc — title, headings,
    opening recitals are enough to name it."""
    try:
        import docx  # python-docx
    except ImportError:
        return ""
    try:
        doc = docx.Document(str(docx_path))
        chunks = []
        for i, para in enumerate(doc.paragraphs):
            if i >= 40:
                break
            text = (para.text or "").strip()
            if text:
                chunks.append(text)
        text = "\n".join(chunks)
        return text[:MAX_CHARS_FOR_PROMPT]
    except Exception as e:
        logger.warning("Failed to read docx %s: %s", docx_path.name, e)
        return ""


def _read_file_head(file_path: Path) -> str:
    """Dispatch to the right reader based on extension. Returns empty string
    if we don't have a reader for this file type (e.g. images)."""
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return _read_pdf_head(file_path)
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx_head(file_path)
    if ext == ".docx":
        return _read_docx_head(file_path)
    return ""


def _clean_suggestion(name: str, ext: str) -> str:
    """Normalize a Claude-proposed filename to match our convention.

    Rules:
      - strip path chars and quotes
      - limit to [\\w.-]
      - collapse repeated separators
      - ensure it ends with the expected extension
    """
    # Remove any path-unsafe characters and quotes
    safe = re.sub(r'[<>:"/\\|?*\']', "", name).strip()
    # Replace spaces with underscores; collapse multiple underscores/hyphens
    safe = re.sub(r"\s+", "_", safe)
    safe = re.sub(r"_+", "_", safe)
    safe = re.sub(r"-+", "-", safe)
    safe = safe.strip("_-. ")

    # Claude may or may not include the extension; enforce it
    if ext and not safe.lower().endswith(ext.lower()):
        safe = re.sub(r"\.[A-Za-z0-9]{1,5}$", "", safe)  # drop any other extension
        safe += ext
    return safe or f"unnamed{ext}"


def _fallback_clean(original_name: str, document_date: Optional[str] = None) -> str:
    """Best-effort cleanup without an LLM call. Used for non-PDFs and when
    the API is unavailable."""
    stem, ext = os.path.splitext(original_name)
    # Drop the YYYY-MM-DD__ prefix the Apps Script adds, if present —
    # we're going to re-add a YYYY-MM prefix.
    stem = re.sub(r"^\d{4}-\d{2}-\d{2}__", "", stem)
    # Drop common noise: docusign IDs in parens, trailing version numbers, GUID-like strings
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
    stem = re.sub(r"_+\(\w+\)\s*$", "", stem)
    stem = re.sub(r"\s+", "_", stem).strip("_- ")
    if not document_date:
        document_date = datetime.now().strftime("%Y-%m")
    return f"{document_date}__{stem}{ext}"


def _prompt_for(filename: str, head_text: str, target_folder: str) -> str:
    return f"""You are naming a document for a small nonprofit's long-term archive.

The document is being filed in: {target_folder}

Current filename: {filename}

First pages of the document (may be truncated):
---
{head_text}
---

Propose ONE concise, descriptive filename using this exact format:
  YYYY-MM__descriptive_topic.ext

Rules:
- YYYY-MM = year and month the document is dated (find it in the content; if unclear, use today's year-month)
- descriptive_topic = snake_case, 3-8 words, no spaces, no special chars, specific to the document's content
- ext = keep the original file extension
- Avoid internal IDs, envelope numbers, Docusign references, version suffixes like (1) (2)

Good examples:
  2026-04__nsia_operating_agreement_amended_restated_executed.pdf
  2026-03__umb_trustee_report_q1_2026.pdf
  2026-02__chase_operating_account_reconciliation.pdf
  2025-12__board_meeting_minutes_december.pdf

Bad examples (don't do this):
  Complete_with_Docusign_North_Shore_Ice_Arena.pdf  (no date, marketing noise)
  NSIA 501C3 Letter.pdf  (no date, vague)
  EXECUTED - NSIA - Amended and Restated Operating Agreement_(33118276_1) (7).pdf  (envelope ID, version suffix)

Return ONLY the filename. No preamble, no markdown, no explanation.
"""


# Short, neutral descriptions of each permanent folder — used by the combined
# destination+name suggester so Claude knows what lives where. Keep these
# tight: one sentence, no NSIA-specific jargon that Claude wouldn't know.
FOLDER_DESCRIPTIONS = {
    "Bond Documents": "Bond indentures, covenants, IFA Series 2008 paperwork, trustee agreements",
    "Financials and Budgets": "Financial statements, P&L, balance sheets, annual financials",
    "Budgets": "Annual and quarterly budget submissions, budget-to-actual workbooks",
    "Bank Statements": "Bank account statements, account reconciliations, account activity reports (Chase, etc)",
    "Board Meetings": "Board agendas, minutes, board memos, resolutions, board packets",
    "Club Sports Consulting Group Contract": "CSCG service contracts, amendments, scope letters",
    "Chiller contract": "Chiller unit service contracts and maintenance agreements",
    "Energy contracts": "Electricity and gas utility bills, energy supply contracts, kWh usage reports",
    "Insurance contract": "Insurance policies, claims, certificates of insurance, renewal paperwork",
    "Ground Leases": "Ground lease documents with Divine Word / Techny property",
    "Operating Agreement": "LLC Operating Agreement (amended/restated) and governance amendments (NOT the Articles of Organization — those go in Corporate Records)",
    "Audit Docs": "Annual audit reports and audit committee correspondence (NOT IRS formation docs — those go in Corporate Records)",
    "NSIA Corporate Records": "Foundational legal/formation documents — Articles of Organization, 501(c)(3) IRS determination letter, bylaws, EIN assignment, annual state reports. One-time records, rarely change.",
    "Ice utilization": "Ice schedules, rink usage reports, utilization breakdowns",
    "2024 Chiller Contract and related matters": "The specific 2024 chiller replacement project",
    "Capital Improvements": "Capex projects, facility upgrades, renovation proposals",
    "Invoices and Bills": "Vendor invoices, bills to pay, remittance advices",
    "Ice resurfacer Quotes": "Zamboni and ice resurfacer bids, specs, quotes",
    "Club Sports Payroll issue": "CSCG-specific payroll correspondence and resolution",
    "Restaurant": "Restaurant and concession operators (Ice Shack, Nick Larkin, F&B term sheets)",
    "Website": "Website development, SEO, domain, hosting contracts",
    "Scoreboard Contract": "Scoreboard sponsorship agreements, equipment contracts",
    "Current Ice contracts": "Active ice rental agreements, league contracts, user group ice agreements",
}


def suggest_destination_and_name(
    file_path: Path,
    candidate_folders: list[str],
    api_key: Optional[str] = None,
) -> tuple[Optional[str], str, str, str]:
    """Read the file and propose BOTH a permanent folder AND a clean filename.

    Returns (folder, filename, source, reasoning):
      folder — one of candidate_folders, or None if Claude couldn't pick
      filename — proposed clean name
      source — "claude" or "fallback"
      reasoning — short explanation (Claude's own words) or fallback message
    """
    original_name = file_path.name
    ext = file_path.suffix

    api_key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None, _fallback_clean(original_name), "fallback", "No ANTHROPIC_API_KEY set — cannot read file content"

    head_text = _read_file_head(file_path)
    if not head_text:
        return None, _fallback_clean(original_name), "fallback", "Couldn't extract text from this file type"

    # Build the folder-choice menu. Only include folders we know about.
    folder_lines = []
    for f in candidate_folders:
        desc = FOLDER_DESCRIPTIONS.get(f, "(no description)")
        folder_lines.append(f"- {f}: {desc}")
    folder_menu = "\n".join(folder_lines)

    prompt = f"""You are filing a document for a small nonprofit's long-term archive (North Shore Ice Arena).

Current filename: {original_name}

First pages / rows of the document:
---
{head_text}
---

Available permanent folders (pick exactly ONE):
{folder_menu}

Your job:
1. Pick the single best folder for this document from the list above. Use the descriptions to guide you.
2. Propose a clean filename in format: YYYY-MM__descriptive_topic.ext
   - YYYY-MM = year and month the document is dated (find it in the content)
   - descriptive_topic = snake_case, 3-8 words, specific to what the document IS
   - ext = same extension as the original ({ext})
3. Give a one-sentence reason for your folder choice.

Return ONLY valid JSON, no markdown fences:
{{
  "folder": "exact folder name from the list above",
  "filename": "YYYY-MM__descriptive_topic{ext}",
  "reasoning": "one short sentence"
}}
"""

    try:
        import anthropic
        import json as _json
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip() if resp.content else ""
        # Strip possible accidental code fences
        if raw.startswith("```"):
            raw = re.sub(r"^```[a-zA-Z]*\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        data = _json.loads(raw)
        folder = data.get("folder")
        if folder not in candidate_folders:
            # Claude hallucinated or abbreviated — reject the folder pick but keep the name
            folder = None
        fn = _clean_suggestion(data.get("filename", ""), ext)
        reasoning = data.get("reasoning", "")
        return folder, fn, "claude", reasoning
    except Exception as e:
        logger.warning("suggest_destination_and_name fell back: %s", e)
        return None, _fallback_clean(original_name), "fallback", f"Claude call failed: {e}"


def suggest_filename(
    file_path: Path,
    target_folder: str,
    api_key: Optional[str] = None,
) -> tuple[str, str]:
    """Propose a clean filename for a file being archived.

    Returns (suggested_name, source) where source is:
      "claude" — LLM read the content and generated a name
      "fallback_clean" — used filename-only heuristics (non-PDF or API unavailable)

    Always returns a usable filename, even when the API call fails — the
    archive flow must not block on naming.
    """
    original_name = file_path.name
    ext = file_path.suffix

    api_key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return _fallback_clean(original_name), "fallback_clean"

    head_text = _read_file_head(file_path)
    if not head_text:
        # Unsupported format (image, binary) or empty document — light cleanup only
        return _fallback_clean(original_name), "fallback_clean"

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            # Haiku is plenty for filename generation and keeps cost ~$0.001/file
            model="claude-haiku-4-5-20251001",
            max_tokens=100,
            messages=[{
                "role": "user",
                "content": _prompt_for(original_name, head_text, target_folder),
            }],
        )
        raw = resp.content[0].text.strip() if resp.content else ""
        if not raw:
            return _fallback_clean(original_name), "fallback_clean"
        return _clean_suggestion(raw, ext), "claude"
    except Exception as e:
        logger.warning("Filename suggester fell back: %s", e)
        return _fallback_clean(original_name), "fallback_clean"
